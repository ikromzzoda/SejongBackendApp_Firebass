import re
import uuid
import random
import secrets
import string
import io
import jwt as pyjwt
from datetime import datetime, timedelta, timezone
from django.conf import settings
from django.contrib.auth.hashers import make_password, check_password
from django.core.mail import send_mail
from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse, OpenApiTypes
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import User, BlacklistedToken, DEFAULT_AVATAR
from groups.models import Group
from utils.jwt_utils import (
    generate_token,
    generate_refresh_token,
    generate_reset_token,
    decode_token,
    RESET_TOKEN_LIFETIME_MINUTES,
)
from utils.decorators import admin_required, jwt_required
from utils.drive import upload_avatar, delete_avatar
from utils.schema import (
    AUTH_HEADER_PARAM,
    ADMIN_RESPONSES,
    UNAUTHORIZED_RESPONSES,
    ErrorResponseSerializer,
    MessageResponseSerializer,
)
from audit_logs.utils import log_action
from .serializers import (
    UserSerializer,
    UserDetailSerializer,
    ProfileSerializer,
    RegisterRequestSerializer,
    RegisterResponseSerializer,
    LoginRequestSerializer,
    LoginResponseSerializer,
    TokenRefreshRequestSerializer,
    TokenRefreshResponseSerializer,
    AdminListUsersResponseSerializer,
    AdminGetUserResponseSerializer,
    AdminEditUserRequestSerializer,
    AdminEditUserResponseSerializer,
    AdminVerifyUserRequestSerializer,
    AdminVerifyUserResponseSerializer,
    AdminSetStatusRequestSerializer,
    AdminSetStatusResponseSerializer,
    AdminCreateUserRequestSerializer,
    AdminCreateUserResponseSerializer,
    UpdateProfileRequestSerializer,
    UpdateProfileResponseSerializer,
    ChangeAvatarRequestSerializer,
    ChangeAvatarResponseSerializer,
    BulkImportRequestSerializer,
    VerifyEmailRequestSerializer,
    ResendEmailCodeRequestSerializer,
    ForgotPasswordRequestSerializer,
    VerifyResetCodeRequestSerializer,
    VerifyResetCodeResponseSerializer,
    ResetPasswordRequestSerializer,
)


PHONE_RE = re.compile(r'^\+992\d{9}$')

ALLOWED_IMAGE_TYPES = {'image/jpeg', 'image/png', 'image/webp'}
MAX_AVATAR_SIZE = 3 * 1024 * 1024  # 3 MB

VALID_STATUSES = {'Student', 'Teacher', 'Admin', 'Guest'}

# Подтверждение email при регистрации — та же политика, что и у сброса пароля
EMAIL_CODE_LIFETIME_MINUTES = 15
EMAIL_CODE_MAX_ATTEMPTS     = 5
EMAIL_CODE_RESEND_SECONDS   = 600  # повторная отправка кода — не чаще раза в 10 минут

# Единый ответ resend-code: не раскрываем, зарегистрирован ли email
RESEND_CODE_MESSAGE = 'Если такой email зарегистрирован и не подтверждён, код отправлен на почту.'


def _clear_email_code(user):
    user.email_code_hash     = ''
    user.email_code_expires  = None
    user.email_code_attempts = 0


def _send_email_verification_code(user, now):
    """Генерирует 6-значный код, сохраняет его хэш на пользователе и шлёт код на почту.
    Поля пользователя только выставляются — save/update вызывает вызывающий код.
    """
    code = f'{secrets.randbelow(1_000_000):06d}'
    user.email_code_hash     = make_password(code)
    user.email_code_expires  = now + timedelta(minutes=EMAIL_CODE_LIFETIME_MINUTES)
    user.email_code_attempts = 0
    user.email_code_sent_at  = now

    try:
        send_mail(
            subject='Sejong: код подтверждения email',
            message=(
                f'Ваш код подтверждения email: {code}\n\n'
                f'Код действует {EMAIL_CODE_LIFETIME_MINUTES} минут.\n'
                'Если вы не регистрировались в Sejong, просто проигнорируйте это письмо.'
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=False,
        )
    except Exception:
        # Не раскрываем наружу проблемы SMTP; код остаётся валидным до истечения
        pass


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------

def _require_fields(data, *fields):
    for field in fields:
        if not data.get(field):
            return Response(
                {'error': f'Поле "{field}" обязательно'},
                status=status.HTTP_400_BAD_REQUEST,
            )
    return None


def _validate_phone(phone: str):
    if not PHONE_RE.match(phone):
        return Response(
            {'error': "Номер должен начинаться с '+992' и содержать 9 цифр после него."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    return None


def _username_taken(username: str) -> bool:
    return bool(list(User.collection.filter('username', '==', username).fetch(1)))


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@extend_schema(
    tags=['Users'],
    summary='Регистрация',
    description=(
        'Регистрация нового пользователя. На указанный email отправляется 6-значный код '
        f'подтверждения (действует {EMAIL_CODE_LIFETIME_MINUTES} минут). До подтверждения '
        'email вход в систему невозможен. Статус по умолчанию — "Guest", ожидает '
        'подтверждения администратора.'
    ),
    request={'multipart/form-data': RegisterRequestSerializer},
    responses={
        201: RegisterResponseSerializer,
        400: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def register(request):
    data = request.data

    err = _require_fields(data, 'username', 'password', 'email', 'phone_number')
    if err:
        return err

    err = _validate_phone(data['phone_number'])
    if err:
        return err

    if _username_taken(data['username']):
        return Response(
            {'error': 'Пользователь с таким username уже существует'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    avatar_file = request.FILES.get('avatar')
    if avatar_file:
        if avatar_file.content_type not in ALLOWED_IMAGE_TYPES:
            return Response(
                {'error': 'Недопустимый формат аватара. Разрешены: JPEG, PNG, WEBP'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if avatar_file.size > MAX_AVATAR_SIZE:
            return Response(
                {'error': 'Файл аватара слишком большой. Максимум 3 МБ'},
                status=status.HTTP_400_BAD_REQUEST,
            )

    user = User()
    user.username      = data['username']
    user.fullname      = data.get('fullname', '')
    user.email         = data['email']
    user.phone_number  = data['phone_number']
    user.password      = make_password(data['password'])
    user.date_of_birth = data.get('date_of_birth', '')
    user.status              = 'Guest'
    user.verification_status = 'Pending'
    user.email_verified       = False
    _send_email_verification_code(user, datetime.now(timezone.utc))
    user.save()

    if avatar_file:
        ext      = avatar_file.name.rsplit('.', 1)[-1] if '.' in avatar_file.name else 'jpg'
        filename = f"avatar_{user.id}.{ext}"
        try:
            new_file_id, new_url = upload_avatar(avatar_file, filename, avatar_file.content_type)
            user.avatar    = new_url
            user.avatar_id = new_file_id
            user.update()
        except Exception:
            pass

    return Response({
        'message': 'Регистрация успешна. Мы отправили 6-значный код на вашу почту — подтвердите email, чтобы войти.',
        'email': user.email,
        'status': user.status,
        'verification_status': user.verification_status,
        'avatar': user.avatar or '',
        'fcm_topic': f'status_{user.status}',
    }, status=status.HTTP_201_CREATED)


@extend_schema(
    tags=['Users'],
    summary='Регистрация — подтвердить email',
    description=(
        'Проверяет 6-значный код из письма, отправленного при регистрации. При верном коде '
        'email считается подтверждённым и пользователь может войти. После '
        f'{EMAIL_CODE_MAX_ATTEMPTS} неверных попыток код аннулируется — нужно запросить новый '
        'через resend-code.'
    ),
    request=VerifyEmailRequestSerializer,
    responses={
        200: MessageResponseSerializer,
        400: ErrorResponseSerializer,
        429: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def verify_email(request):
    email = (request.data.get('email') or '').strip()
    code  = (request.data.get('code') or '').strip()
    if not email or not code:
        return Response({'error': 'Поля "email" и "code" обязательны'}, status=status.HTTP_400_BAD_REQUEST)

    users = list(User.collection.filter('email', '==', email).fetch(1))
    user  = users[0] if users else None

    if user and user.email_verified:
        return Response({'message': 'Email уже подтверждён. Можете войти.'})

    if not user or not user.email_code_hash:
        return Response(
            {'error': 'Код неверен или истёк. Запросите новый код.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    now = datetime.now(timezone.utc)
    if not user.email_code_expires or now > user.email_code_expires:
        _clear_email_code(user)
        user.update()
        return Response(
            {'error': 'Код истёк. Запросите новый код.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not check_password(code, user.email_code_hash):
        attempts = (user.email_code_attempts or 0) + 1
        if attempts >= EMAIL_CODE_MAX_ATTEMPTS:
            _clear_email_code(user)
            user.update()
            return Response(
                {'error': 'Слишком много неверных попыток. Запросите новый код.'},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
            )
        user.email_code_attempts = attempts
        user.update()
        return Response(
            {'error': f'Неверный код. Осталось попыток: {EMAIL_CODE_MAX_ATTEMPTS - attempts}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Код верный: подтверждаем email и аннулируем код
    user.email_verified = True
    _clear_email_code(user)
    user.update()

    return Response({'message': 'Email подтверждён. Теперь вы можете войти.'})


@extend_schema(
    tags=['Users'],
    summary='Регистрация — отправить код повторно',
    description=(
        'Повторно отправляет 6-значный код подтверждения email. Повторная отправка — не чаще '
        f'раза в {EMAIL_CODE_RESEND_SECONDS // 60} минут. Ответ одинаковый независимо от того, '
        'существует ли email (защита от перебора).'
    ),
    request=ResendEmailCodeRequestSerializer,
    responses={
        200: MessageResponseSerializer,
        400: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def resend_email_code(request):
    email = (request.data.get('email') or '').strip()
    if not email:
        return Response({'error': 'Поле "email" обязательно'}, status=status.HTTP_400_BAD_REQUEST)

    users = list(User.collection.filter('email', '==', email).fetch(1))
    if not users or users[0].email_verified:
        return Response({'message': RESEND_CODE_MESSAGE})

    user = users[0]
    now  = datetime.now(timezone.utc)

    # Rate limit: не отправляем новый код, если предыдущий ушёл недавно
    if user.email_code_sent_at and (now - user.email_code_sent_at).total_seconds() < EMAIL_CODE_RESEND_SECONDS:
        return Response({'message': RESEND_CODE_MESSAGE})

    _send_email_verification_code(user, now)
    user.update()

    return Response({'message': RESEND_CODE_MESSAGE})


@extend_schema(
    tags=['Users'],
    summary='Вход',
    description='Вход по username и паролю. Возвращает access и refresh токены.',
    request=LoginRequestSerializer,
    responses={
        200: LoginResponseSerializer,
        400: ErrorResponseSerializer,
        401: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def login(request):
    data     = request.data
    username = data.get('username', '').strip()
    password = data.get('password', '')

    device_token = data.get('device_token', '').strip()

    if not username or not password or not device_token:
        return Response(
            {'error': 'Введите username, пароль и device_token'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    users = list(User.collection.filter('username', '==', username).fetch(1))
    if not users or not check_password(password, users[0].password):
        return Response(
            {'error': 'Неверный username или пароль'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    user = users[0]

    # email_verified is False — только у новых пользователей, не подтвердивших код;
    # у старых записей поле отсутствует (None) — их не блокируем
    if user.email_verified is False:
        return Response(
            {'error': 'Email не подтверждён. Введите код из письма или запросите новый.'},
            status=status.HTTP_403_FORBIDDEN,
        )

    if device_token != (user.device_token or ''):
        user.device_token = device_token

    access_token = generate_token(
        user_id=user.id,
        username=user.username,
        status=user.status,
        verification_status=user.verification_status,
    )
    refresh_token, refresh_jti = generate_refresh_token(user.id)
    user.refresh_token_jti = refresh_jti
    user.update()

    return Response({
        'message':       'Вход выполнен',
        'token':         access_token,
        'refresh_token': refresh_token,
        'status':        user.status,
        'verification_status': user.verification_status,
    })


@extend_schema(
    tags=['Users'],
    summary='Выход',
    description='Отзывает текущий access-токен (добавляет его jti в чёрный список) и очищает refresh-токен пользователя.',
    parameters=[AUTH_HEADER_PARAM],
    request=None,
    responses={
        200: MessageResponseSerializer,
        401: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def logout(request):
    auth_header = request.headers.get('Authorization', '')
    if not auth_header.startswith('Bearer '):
        return Response({'error': 'Токен не предоставлен'}, status=status.HTTP_401_UNAUTHORIZED)

    token = auth_header.split(' ', 1)[1]
    try:
        payload = decode_token(token)
    except pyjwt.ExpiredSignatureError:
        return Response({'error': 'Токен уже истёк'}, status=status.HTTP_401_UNAUTHORIZED)
    except pyjwt.InvalidTokenError:
        return Response({'error': 'Недействительный токен'}, status=status.HTTP_401_UNAUTHORIZED)

    bt = BlacklistedToken()
    bt.id = payload['jti']
    bt.save()

    try:
        user = User.collection.get(f"users/{payload['user_id']}")
        if user:
            user.refresh_token_jti = ''
            user.update()
    except Exception:
        pass

    return Response({'message': 'Выход выполнен'})


@extend_schema(
    tags=['Users'],
    summary='Обновить токен',
    description='Обновляет access-токен по действительному refresh-токену (одноразовый — старый refresh-токен становится недействителен).',
    request=TokenRefreshRequestSerializer,
    responses={
        200: TokenRefreshResponseSerializer,
        400: ErrorResponseSerializer,
        401: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def token_refresh(request):
    incoming = request.data.get('refresh_token', '')
    if not incoming:
        return Response({'error': 'refresh_token не передан'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        payload = decode_token(incoming)
    except pyjwt.ExpiredSignatureError:
        return Response({'error': 'Refresh token истёк. Войдите заново.'}, status=status.HTTP_401_UNAUTHORIZED)
    except pyjwt.InvalidTokenError:
        return Response({'error': 'Недействительный refresh token'}, status=status.HTTP_401_UNAUTHORIZED)

    if payload.get('type') != 'refresh':
        return Response({'error': 'Недействительный refresh token'}, status=status.HTTP_401_UNAUTHORIZED)

    try:
        user = User.collection.get(f"users/{payload['user_id']}")
    except Exception:
        user = None
    if not user or user.refresh_token_jti != payload['jti']:
        return Response({'error': 'Refresh token недействителен или уже использован'}, status=status.HTTP_401_UNAUTHORIZED)

    new_access = generate_token(
        user_id=user.id,
        username=user.username,
        status=user.status,
        verification_status=user.verification_status,
    )
    new_refresh, new_jti = generate_refresh_token(user.id)
    user.refresh_token_jti = new_jti
    user.update()

    return Response({
        'token':         new_access,
        'refresh_token': new_refresh,
    })


# ---------------------------------------------------------------------------
# Password reset (забыли пароль)
# ---------------------------------------------------------------------------

RESET_CODE_LIFETIME_MINUTES = 15
RESET_CODE_MAX_ATTEMPTS     = 5
RESET_CODE_RESEND_SECONDS   = 600  # повторная отправка кода — не чаще раза в 10 минут

# Единый ответ forgot-password: не раскрываем, зарегистрирован ли email
FORGOT_PASSWORD_MESSAGE = 'Если такой email зарегистрирован, код отправлен на почту.'


def _clear_reset_code(user):
    user.reset_code_hash     = ''
    user.reset_code_expires  = None
    user.reset_code_attempts = 0


@extend_schema(
    tags=['Users'],
    summary='Забыли пароль — отправить код',
    description=(
        'Отправляет 6-значный код на email пользователя. Код действует '
        f'{RESET_CODE_LIFETIME_MINUTES} минут, повторная отправка — не чаще раза в '
        f'{RESET_CODE_RESEND_SECONDS // 60} минут. Ответ одинаковый независимо от того, '
        'существует ли email (защита от перебора).'
    ),
    request=ForgotPasswordRequestSerializer,
    responses={
        200: MessageResponseSerializer,
        400: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def forgot_password(request):
    email = (request.data.get('email') or '').strip()
    if not email:
        return Response({'error': 'Поле "email" обязательно'}, status=status.HTTP_400_BAD_REQUEST)

    users = list(User.collection.filter('email', '==', email).fetch(1))
    if not users:
        return Response({'message': FORGOT_PASSWORD_MESSAGE})

    user = users[0]
    now  = datetime.now(timezone.utc)

    # Rate limit: не отправляем новый код, если предыдущий ушёл меньше минуты назад
    if user.reset_code_sent_at and (now - user.reset_code_sent_at).total_seconds() < RESET_CODE_RESEND_SECONDS:
        return Response({'message': FORGOT_PASSWORD_MESSAGE})

    code = f'{secrets.randbelow(1_000_000):06d}'
    user.reset_code_hash     = make_password(code)
    user.reset_code_expires  = now + timedelta(minutes=RESET_CODE_LIFETIME_MINUTES)
    user.reset_code_attempts = 0
    user.reset_code_sent_at  = now
    user.update()

    try:
        send_mail(
            subject='Sejong: код для сброса пароля',
            message=(
                f'Ваш код для сброса пароля: {code}\n\n'
                f'Код действует {RESET_CODE_LIFETIME_MINUTES} минут.\n'
                'Если вы не запрашивали сброс пароля, просто проигнорируйте это письмо.'
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False,
        )
    except Exception:
        # Не раскрываем наружу проблемы SMTP; код остаётся валидным до истечения
        pass

    return Response({'message': FORGOT_PASSWORD_MESSAGE})


@extend_schema(
    tags=['Users'],
    summary='Забыли пароль — проверить код',
    description=(
        'Проверяет 6-значный код из письма. При верном коде возвращает короткоживущий '
        'reset_token для смены пароля. После '
        f'{RESET_CODE_MAX_ATTEMPTS} неверных попыток код аннулируется — нужно запросить новый.'
    ),
    request=VerifyResetCodeRequestSerializer,
    responses={
        200: VerifyResetCodeResponseSerializer,
        400: ErrorResponseSerializer,
        429: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def verify_reset_code(request):
    email = (request.data.get('email') or '').strip()
    code  = (request.data.get('code') or '').strip()
    if not email or not code:
        return Response({'error': 'Поля "email" и "code" обязательны'}, status=status.HTTP_400_BAD_REQUEST)

    users = list(User.collection.filter('email', '==', email).fetch(1))
    user  = users[0] if users else None

    if not user or not user.reset_code_hash:
        return Response(
            {'error': 'Код неверен или истёк. Запросите новый код.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    now = datetime.now(timezone.utc)
    if not user.reset_code_expires or now > user.reset_code_expires:
        _clear_reset_code(user)
        user.update()
        return Response(
            {'error': 'Код истёк. Запросите новый код.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not check_password(code, user.reset_code_hash):
        attempts = (user.reset_code_attempts or 0) + 1
        if attempts >= RESET_CODE_MAX_ATTEMPTS:
            _clear_reset_code(user)
            user.update()
            return Response(
                {'error': 'Слишком много неверных попыток. Запросите новый код.'},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
            )
        user.reset_code_attempts = attempts
        user.update()
        return Response(
            {'error': f'Неверный код. Осталось попыток: {RESET_CODE_MAX_ATTEMPTS - attempts}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Код верный: аннулируем его и выдаём одноразовый reset-токен
    reset_token, reset_jti = generate_reset_token(user.id)
    _clear_reset_code(user)
    user.reset_token_jti = reset_jti
    user.update()

    return Response({
        'message': 'Код подтверждён. Установите новый пароль.',
        'reset_token': reset_token,
    })


@extend_schema(
    tags=['Users'],
    summary='Забыли пароль — установить новый',
    description=(
        'Устанавливает новый пароль по reset_token из verify-code. Токен одноразовый и '
        f'действует {RESET_TOKEN_LIFETIME_MINUTES} минут. После смены пароля все '
        'refresh-токены пользователя отзываются.'
    ),
    request=ResetPasswordRequestSerializer,
    responses={
        200: MessageResponseSerializer,
        400: ErrorResponseSerializer,
        401: ErrorResponseSerializer,
    },
)
@api_view(['POST'])
def reset_password(request):
    reset_token  = request.data.get('reset_token', '')
    new_password = request.data.get('new_password', '')

    if not reset_token or not new_password:
        return Response(
            {'error': 'Поля "reset_token" и "new_password" обязательны'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if len(new_password) < 6:
        return Response(
            {'error': 'Пароль должен содержать минимум 6 символов'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        payload = decode_token(reset_token)
    except pyjwt.ExpiredSignatureError:
        return Response(
            {'error': 'Срок действия reset_token истёк. Запросите код заново.'},
            status=status.HTTP_401_UNAUTHORIZED,
        )
    except pyjwt.InvalidTokenError:
        return Response({'error': 'Недействительный reset_token'}, status=status.HTTP_401_UNAUTHORIZED)

    if payload.get('type') != 'reset':
        return Response({'error': 'Недействительный reset_token'}, status=status.HTTP_401_UNAUTHORIZED)

    try:
        user = User.collection.get(f"users/{payload['user_id']}")
    except Exception:
        user = None
    if not user or user.reset_token_jti != payload['jti']:
        return Response(
            {'error': 'Reset token недействителен или уже использован'},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    user.password        = make_password(new_password)
    user.reset_token_jti = ''
    # Отзываем refresh-токены: после сброса пароля старые сессии недействительны
    user.refresh_token_jti = ''
    user.update()

    return Response({'message': 'Пароль изменён. Войдите с новым паролем.'})


# ---------------------------------------------------------------------------
# Profile — get my data
# ---------------------------------------------------------------------------

@extend_schema(
    tags=['Users'],
    summary='Мои данные',
    description='Возвращает данные профиля текущего авторизованного пользователя.',
    parameters=[AUTH_HEADER_PARAM],
    responses={
        200: ProfileSerializer,
        404: ErrorResponseSerializer,
        **UNAUTHORIZED_RESPONSES,
    },
)
@api_view(['GET'])
@jwt_required
def get_profile(request):
    user_id = request.user_payload.get('user_id', '')
    try:
        user = User.collection.get(f'users/{user_id}')
    except Exception:
        user = None
    if not user:
        return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)

    return Response({
        'id':                  user.id,
        'username':            user.username or '',
        'fullname':            user.fullname or '',
        'email':               user.email or '',
        'phone_number':        user.phone_number or '',
        'date_of_birth':       user.date_of_birth or '',
        'status':              user.status or '',
        'verification_status': user.verification_status or '',
        'group_id':            user.group or '',
        'avatar':              user.avatar or '',
        'date_joined':         str(user.date_joined) if user.date_joined else '',
    })


# ---------------------------------------------------------------------------
# Admin — helpers
# ---------------------------------------------------------------------------

def _resolve_group_name(group_id: str, cache: dict | None = None) -> str:
    if not group_id:
        return ''
    if cache is not None:
        return cache.get(group_id, group_id)
    try:
        g = Group.collection.get(f'groups/{group_id}')
        return g.name if g else group_id
    except Exception:
        return group_id


def _user_dict(user, groups_cache: dict | None = None, full: bool = False):
    d = {
        'id':                  user.id,
        'username':            user.username,
        'fullname':            user.fullname,
        'email':               user.email,
        'phone_number':        user.phone_number,
        'status':              user.status,
        'verification_status': user.verification_status,
        'group_id':            user.group or '',
        'group':               _resolve_group_name(user.group or '', groups_cache),
        'avatar':              user.avatar or '',
        'date_joined':         str(user.date_joined) if user.date_joined else '',
    }
    if full:
        d['date_of_birth'] = user.date_of_birth or ''
        d['device_token']  = user.device_token or ''
    return d


# ---------------------------------------------------------------------------
# Admin — users
# ---------------------------------------------------------------------------

@extend_schema(
    tags=['Users'],
    operation_id='users_admin_list_users',
    summary='Список пользователей (admin)',
    description='Список всех пользователей с опциональной фильтрацией по статусу, статусу верификации или группе.',
    parameters=[
        AUTH_HEADER_PARAM,
        OpenApiParameter('status', OpenApiTypes.STR, description='Фильтр по статусу (Student/Teacher/Admin/Guest)'),
        OpenApiParameter('verification_status', OpenApiTypes.STR, description='Фильтр по статусу верификации (Pending/Approved/Rejected)'),
        OpenApiParameter('group_id', OpenApiTypes.STR, description='Фильтр по ID группы'),
    ],
    responses={
        200: AdminListUsersResponseSerializer,
        **ADMIN_RESPONSES,
    },
)
@api_view(['GET'])
@admin_required
def admin_list_users(request):
    """Список всех пользователей с фильтрацией.
    Query params: ?status=  ?verification_status=  ?group_id=
    """
    filter_status = request.query_params.get('status')
    filter_verify = request.query_params.get('verification_status')
    filter_group  = request.query_params.get('group_id')

    if filter_status and filter_status in VALID_STATUSES:
        users = list(User.collection.filter('status', '==', filter_status).fetch(500))
    elif filter_verify and filter_verify in ('Pending', 'Approved', 'Rejected'):
        users = list(User.collection.filter('verification_status', '==', filter_verify).fetch(500))
    elif filter_group:
        users = list(User.collection.filter('group', '==', filter_group).fetch(500))
    else:
        users = list(User.collection.fetch(500))

    groups_cache = {g.id: g.name for g in Group.collection.fetch(100)}

    return Response({
        'total': len(users),
        'users': [_user_dict(u, groups_cache) for u in users],
    })


@extend_schema(
    tags=['Users'],
    summary='Получить пользователя (admin)',
    description='Получить полную информацию об одном пользователе по ID.',
    parameters=[AUTH_HEADER_PARAM],
    responses={
        200: AdminGetUserResponseSerializer,
        404: ErrorResponseSerializer,
        **ADMIN_RESPONSES,
    },
)
@api_view(['GET'])
@admin_required
def admin_get_user(request, user_id):
    """Получить одного пользователя по ID."""
    try:
        user = User.collection.get(f'users/{user_id}')
    except Exception:
        user = None
    if not user:
        return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)
    return Response({'user': _user_dict(user, full=True)})


@extend_schema(
    tags=['Users'],
    summary='Редактировать пользователя (admin)',
    description='Частичное обновление данных пользователя. Передавайте только изменяемые поля.',
    parameters=[AUTH_HEADER_PARAM],
    request=AdminEditUserRequestSerializer,
    responses={
        200: AdminEditUserResponseSerializer,
        400: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        **ADMIN_RESPONSES,
    },
)
@api_view(['PATCH'])
@admin_required
def admin_edit_user(request, user_id):
    """Редактировать данные пользователя."""
    try:
        user = User.collection.get(f'users/{user_id}')
    except Exception:
        user = None
    if not user:
        return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)

    data = request.data
    updated_fields = []

    if 'fullname' in data:
        user.fullname = data['fullname']
        updated_fields.append('fullname')

    if 'email' in data:
        user.email = data['email']
        updated_fields.append('email')

    if 'phone_number' in data:
        phone = data['phone_number'].strip()
        err = _validate_phone(phone)
        if err:
            return err
        user.phone_number = phone
        updated_fields.append('phone_number')

    if 'date_of_birth' in data:
        user.date_of_birth = data['date_of_birth']
        updated_fields.append('date_of_birth')

    if 'status' in data:
        if data['status'] not in VALID_STATUSES:
            return Response(
                {'error': f'Допустимые статусы: {", ".join(VALID_STATUSES)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user.status = data['status']
        updated_fields.append('status')

    if 'verification_status' in data:
        if data['verification_status'] not in ('Pending', 'Approved', 'Rejected'):
            return Response(
                {'error': 'Допустимые значения: Pending, Approved, Rejected'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        user.verification_status = data['verification_status']
        updated_fields.append('verification_status')

    if 'group_id' in data:
        group_id = data['group_id']
        if group_id:
            try:
                group = Group.collection.get(f'groups/{group_id}')
            except Exception:
                group = None
            if not group:
                return Response({'error': 'Группа не найдена'}, status=status.HTTP_404_NOT_FOUND)
        user.group = group_id
        updated_fields.append('group')

    if 'group' in data and 'group_id' not in data:
        group_name = data['group'].strip()
        if group_name:
            found = list(Group.collection.filter('name', '==', group_name).fetch(1))
            if not found:
                return Response({'error': f'Группа "{group_name}" не найдена'}, status=status.HTTP_404_NOT_FOUND)
            user.group = found[0].id
        else:
            user.group = ''
        updated_fields.append('group')

    if 'password' in data:
        new_pass = data['password']
        if not new_pass:
            return Response({'error': 'Пароль не может быть пустым'}, status=status.HTTP_400_BAD_REQUEST)
        user.password = make_password(new_pass)
        updated_fields.append('password')

    if not updated_fields:
        return Response({'message': 'Нет данных для обновления'}, status=status.HTTP_400_BAD_REQUEST)

    user.update()
    log_action(request, 'update', 'User', user_id, {'updated_fields': updated_fields})
    return Response({
        'message': 'Данные пользователя обновлены',
        'updated_fields': updated_fields,
        'user': _user_dict(user),
    })


@extend_schema(
    tags=['Users'],
    summary='Подтвердить/отклонить верификацию (admin)',
    description='При подтверждении статус пользователя меняется на "Student".',
    parameters=[AUTH_HEADER_PARAM],
    request=AdminVerifyUserRequestSerializer,
    responses={
        200: AdminVerifyUserResponseSerializer,
        400: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        **ADMIN_RESPONSES,
    },
)
@api_view(['POST'])
@admin_required
def admin_verify_user(request, user_id):
    """Подтвердить или отклонить верификацию.
    Body: { "action": "approve" | "reject" }
    """
    action = request.data.get('action')
    if action not in ('approve', 'reject'):
        return Response(
            {'error': 'Поле "action" должно быть "approve" или "reject"'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        user = User.collection.get(f'users/{user_id}')
    except Exception:
        user = None
    if not user:
        return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)

    if action == 'approve':
        user.verification_status = 'Approved'
        user.status = 'Student'
    else:
        user.verification_status = 'Rejected'

    user.update()
    log_action(request, 'update', 'User', user_id, {
        'verification_action': action,
        'verification_status': user.verification_status,
    })
    return Response({
        'message': f'Пользователь {"подтверждён" if action == "approve" else "отклонён"}.',
        'user': _user_dict(user),
    })


@extend_schema(
    tags=['Users'],
    summary='Назначить статус (admin)',
    parameters=[AUTH_HEADER_PARAM],
    request=AdminSetStatusRequestSerializer,
    responses={
        200: AdminSetStatusResponseSerializer,
        400: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        **ADMIN_RESPONSES,
    },
)
@api_view(['POST'])
@admin_required
def admin_set_status(request, user_id):
    """Назначить статус пользователю.
    Body: { "status": "Student" | "Teacher" | "Admin" | "Guest" }
    """
    new_status = request.data.get('status')
    if new_status not in VALID_STATUSES:
        return Response(
            {'error': f'Допустимые статусы: {", ".join(VALID_STATUSES)}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        user = User.collection.get(f'users/{user_id}')
    except Exception:
        user = None
    if not user:
        return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)

    user.status = new_status
    user.update()
    log_action(request, 'update', 'User', user_id, {'new_status': new_status})
    return Response({
        'message': f'Статус пользователя изменён на "{new_status}".',
        'user': _user_dict(user),
    })


@extend_schema(
    tags=['Users'],
    summary='Создать пользователя (admin)',
    description='Создать нового пользователя вручную. Верификация выставляется сразу как "Approved".',
    parameters=[AUTH_HEADER_PARAM],
    request=AdminCreateUserRequestSerializer,
    responses={
        201: AdminCreateUserResponseSerializer,
        400: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        **ADMIN_RESPONSES,
    },
)
@api_view(['POST'])
@admin_required
def admin_create_user(request):
    """Создать нового пользователя вручную."""
    data = request.data

    err = _require_fields(data, 'username', 'password', 'email', 'phone_number')
    if err:
        return err

    err = _validate_phone(data['phone_number'])
    if err:
        return err

    if _username_taken(data['username']):
        return Response(
            {'error': 'Пользователь с таким username уже существует'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user_status = data.get('status', 'Student')
    if user_status not in VALID_STATUSES:
        return Response(
            {'error': f'Допустимые статусы: {", ".join(VALID_STATUSES)}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    group_id   = data.get('group_id', '')
    group_name = data.get('group', '')
    if group_name and not group_id:
        found = list(Group.collection.filter('name', '==', group_name).fetch(1))
        if not found:
            return Response({'error': f'Группа "{group_name}" не найдена'}, status=status.HTTP_404_NOT_FOUND)
        group_id = found[0].id
    elif group_id:
        try:
            group = Group.collection.get(f'groups/{group_id}')
        except Exception:
            group = None
        if not group:
            return Response({'error': 'Группа не найдена'}, status=status.HTTP_404_NOT_FOUND)

    user = User()
    user.username      = data['username']
    user.password      = make_password(data['password'])
    user.fullname      = data.get('fullname', '')
    user.email         = data['email']
    user.phone_number  = data['phone_number']
    user.date_of_birth = data.get('date_of_birth', '')
    user.status        = user_status
    user.group         = group_id
    user.avatar        = data.get('avatar', DEFAULT_AVATAR)
    user.verification_status = 'Approved'
    user.email_verified       = True
    user.save()
    log_action(request, 'create', 'User', user.id, {
        'username': user.username,
        'status':   user.status,
    })

    return Response({
        'message': f'Пользователь "{user.username}" успешно создан.',
        'user': _user_dict(user),
    }, status=status.HTTP_201_CREATED)


# ---------------------------------------------------------------------------
# Profile
# ---------------------------------------------------------------------------

@extend_schema(
    tags=['Users'],
    summary='Обновить профиль',
    description='Частичное обновление своего профиля. Для смены пароля обязательно передать текущий пароль в поле "check_password".',
    parameters=[AUTH_HEADER_PARAM],
    request=UpdateProfileRequestSerializer,
    responses={
        200: UpdateProfileResponseSerializer,
        400: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        **UNAUTHORIZED_RESPONSES,
    },
)
@api_view(['PATCH'])
@jwt_required
def update_profile(request):
    """Обновить данные своего профиля."""
    data    = request.data
    payload = request.user_payload
    user_id = payload['user_id']

    try:
        user = User.collection.get(f'users/{user_id}')
    except Exception:
        user = None
    if not user:
        return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)

    updated_fields = []

    if 'username' in data:
        new_username = data['username'].strip()
        if not new_username:
            return Response({'error': 'Username не может быть пустым'}, status=status.HTTP_400_BAD_REQUEST)
        if new_username != user.username:
            if _username_taken(new_username):
                return Response({'error': 'Пользователь с таким username уже существует'}, status=status.HTTP_400_BAD_REQUEST)
            user.username = new_username
            updated_fields.append('username')

    if 'email' in data:
        user.email = data['email'].strip()
        updated_fields.append('email')

    if 'phone_number' in data:
        phone = data['phone_number'].strip()
        err = _validate_phone(phone)
        if err:
            return err
        user.phone_number = phone
        updated_fields.append('phone_number')

    if 'password' in data:
        current = data.get('check_password', '')
        if not current:
            return Response(
                {'error': "Для смены пароля укажите текущий пароль в поле 'check_password'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not check_password(current, user.password):
            return Response({'error': 'Неверный текущий пароль'}, status=status.HTTP_400_BAD_REQUEST)
        user.password = make_password(data['password'])
        updated_fields.append('password')

    if not updated_fields:
        return Response({'message': 'Нет данных для обновления'}, status=status.HTTP_400_BAD_REQUEST)

    user.update()

    new_token = generate_token(
        user_id=user.id,
        username=user.username,
        status=user.status,
        verification_status=user.verification_status,
    )

    return Response({
        'message': 'Профиль обновлён',
        'updated_fields': updated_fields,
        'token': new_token,
    })


@extend_schema(
    tags=['Users'],
    summary='Сменить аватар',
    parameters=[AUTH_HEADER_PARAM],
    request={'multipart/form-data': ChangeAvatarRequestSerializer},
    responses={
        200: ChangeAvatarResponseSerializer,
        400: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        502: ErrorResponseSerializer,
        **UNAUTHORIZED_RESPONSES,
    },
)
@api_view(['POST'])
@jwt_required
def change_avatar(request):
    """Заменить аватар пользователя (multipart/form-data, поле "avatar")."""
    file = request.FILES.get('avatar')
    if not file:
        return Response({'error': 'Файл не передан. Используйте поле "avatar"'}, status=status.HTTP_400_BAD_REQUEST)

    if file.content_type not in ALLOWED_IMAGE_TYPES:
        return Response(
            {'error': 'Недопустимый формат. Разрешены: JPEG, PNG, WEBP'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if file.size > MAX_AVATAR_SIZE:
        return Response({'error': 'Файл слишком большой. Максимум 3 МБ'}, status=status.HTTP_400_BAD_REQUEST)

    user_id = request.user_payload['user_id']
    try:
        user = User.collection.get(f'users/{user_id}')
    except Exception:
        user = None
    if not user:
        return Response({'error': 'Пользователь не найден'}, status=status.HTTP_404_NOT_FOUND)

    if user.avatar_id:
        try:
            delete_avatar(user.avatar_id)
        except Exception:
            pass

    ext = file.name.rsplit('.', 1)[-1] if '.' in file.name else 'jpg'
    filename = f"avatar_{user_id}.{ext}"

    try:
        new_file_id, new_url = upload_avatar(file, filename, file.content_type)
    except Exception as e:
        return Response({'error': f'Ошибка загрузки на Google Drive: {str(e)}'}, status=status.HTTP_502_BAD_GATEWAY)

    user.avatar    = new_url
    user.avatar_id = new_file_id
    user.update()

    return Response({
        'message': 'Аватар успешно обновлён',
        'avatar': new_url,
    })


# ---------------------------------------------------------------------------
# Bulk import
# ---------------------------------------------------------------------------

_COL_ALIASES = {
    'fullname':      ['full name', 'фио', 'ф.и.о', 'ф.и.о.', 'имя', 'полное имя', 'name', 'имя фамилия'],
    'email':         ['email', 'почта', 'e-mail', 'эл. почта', 'электронная почта'],
    'phone_number':  ['phone number', 'phone_number', 'телефон', 'номер', 'номер телефона', 'моб', 'моб.', 'тел'],
    'date_of_birth': ['date of birth/생년월일', 'дата рождения', 'дата', 'birth', 'д.р.', 'день рождения'],
    'group':         ['group', 'группа', 'учебная группа', 'класс'],
}

_TRANSLIT = {
    'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'yo','ж':'zh','з':'z',
    'и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r',
    'с':'s','т':'t','у':'u','ф':'f','х':'kh','ц':'ts','ч':'ch','ш':'sh','щ':'sch',
    'ъ':'','ы':'y','ь':'','э':'e','ю':'yu','я':'ya',
}


def _transliterate(text: str) -> str:
    return ''.join(_TRANSLIT.get(c, c) for c in text.lower())


def _generate_username(fullname: str, batch_taken: set) -> str:
    parts = fullname.strip().split()
    if parts:
        base = _transliterate(parts[0])
        base = re.sub(r'[^a-z0-9]', '', base)[:12] or 'student'
    else:
        base = 'student'
    for _ in range(20):
        candidate = f"{base}_{random.randint(1000, 9999)}"
        if candidate not in batch_taken and not _username_taken(candidate):
            return candidate
    return f"student_{uuid.uuid4().hex[:8]}"


def _generate_password(length: int = 10) -> str:
    chars = string.ascii_letters + string.digits
    pwd = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
    ]
    pwd += [secrets.choice(chars) for _ in range(length - 3)]
    secrets.SystemRandom().shuffle(pwd)
    return ''.join(pwd)


def _detect_columns(header_row) -> dict:
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell.value is None:
            continue
        normalized = str(cell.value).strip().lower()
        for field, aliases in _COL_ALIASES.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = idx
    return mapping


def _style_header(ws, col_count: int):
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


@extend_schema(
    tags=['Users'],
    summary='Массовая загрузка студентов (admin)',
    description='Загружает студентов из Excel-файла (.xlsx) и возвращает файл .xlsx с результатами импорта (логины/пароли/ошибки).',
    parameters=[AUTH_HEADER_PARAM],
    request={'multipart/form-data': BulkImportRequestSerializer},
    responses={
        200: OpenApiResponse(
            response=OpenApiTypes.BINARY,
            description='Файл .xlsx с результатами импорта (students_credentials.xlsx).',
        ),
        400: ErrorResponseSerializer,
        **ADMIN_RESPONSES,
    },
)
@api_view(['POST'])
@admin_required
def admin_bulk_import(request):
    """Массовая загрузка студентов из Excel (.xlsx), поле "file"."""
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'Файл не передан. Используйте поле "file"'}, status=status.HTTP_400_BAD_REQUEST)

    if not excel_file.name.endswith('.xlsx'):
        return Response({'error': 'Разрешён только формат .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb_in = load_workbook(excel_file, data_only=True)
    except Exception:
        return Response({'error': 'Не удалось открыть файл. Убедитесь что это корректный .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

    ws_in = wb_in.active
    rows  = list(ws_in.iter_rows())
    if len(rows) < 2:
        return Response({'error': 'Файл пустой или содержит только заголовок'}, status=status.HTTP_400_BAD_REQUEST)

    col_map = _detect_columns(rows[0])
    if not col_map:
        return Response(
            {'error': 'Не найдены известные заголовки. Ожидаются: ФИО, Email, Телефон, Группа, Дата рождения'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    batch_usernames: set = set()
    groups_by_name  = {g.name.lower(): g.id for g in Group.collection.fetch(100)}

    results = []
    for row in rows[1:]:
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue

        def get(field):
            idx = col_map.get(field)
            return str(values[idx]).strip() if idx is not None and values[idx] is not None else ''

        fullname      = get('fullname')
        email         = get('email')
        phone_number  = re.sub(r'[^\d+]', '', get('phone_number'))
        date_of_birth = get('date_of_birth')
        group_name    = get('group')

        username = _generate_username(fullname or 'student', batch_usernames)
        password = _generate_password()
        batch_usernames.add(username)

        group_id = groups_by_name.get(group_name.lower(), '') if group_name else ''

        try:
            user = User()
            user.username      = username
            user.password      = make_password(password)
            user.fullname      = fullname
            user.email         = email
            user.phone_number  = phone_number
            user.date_of_birth = date_of_birth
            user.status        = 'Student'
            user.group         = group_id
            user.avatar        = DEFAULT_AVATAR
            user.verification_status = 'Approved'
            user.email_verified       = True
            user.save()
            results.append((fullname, email, phone_number, group_name, username, password, 'Успешно', ''))
        except Exception as e:
            results.append((fullname, email, phone_number, group_name, '', '', 'Ошибка', str(e)))

    wb_out  = Workbook()
    ws_out  = wb_out.active
    ws_out.title = 'Результаты импорта'
    ws_out.row_dimensions[1].height = 20

    headers = ['№', 'ФИО', 'Email', 'Телефон', 'Группа', 'Username', 'Password', 'Статус', 'Примечание']
    ws_out.append(headers)
    _style_header(ws_out, len(headers))

    green_fill = PatternFill('solid', fgColor='E2EFDA')
    red_fill   = PatternFill('solid', fgColor='FFDDC1')

    for i, (fullname, email, phone, group, username, password, status_text, note) in enumerate(results, start=1):
        ws_out.append([i, fullname, email, phone, group, username, password, status_text, note])
        row_fill = green_fill if status_text == 'Успешно' else red_fill
        for col in range(1, len(headers) + 1):
            ws_out.cell(row=i + 1, column=col).fill      = row_fill
            ws_out.cell(row=i + 1, column=col).alignment = Alignment(vertical='center')

    for row in ws_out.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.font = Font(bold=True)

    _auto_width(ws_out)

    success_count = sum(1 for r in results if r[6] == 'Успешно')
    error_count   = len(results) - success_count
    log_action(request, 'create', 'User', 'bulk_import', {
        'total':   len(results),
        'success': success_count,
        'errors':  error_count,
    })

    ws_out.append([])
    ws_out.append(['', f'Итого: {len(results)} строк | Успешно: {success_count} | Ошибок: {error_count}'])
    ws_out.cell(row=ws_out.max_row, column=2).font = Font(bold=True, size=11)

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="students_credentials.xlsx"'
    return response


@extend_schema(
    tags=['Users'],
    summary='Шаблон Excel для импорта (admin)',
    description='Скачать шаблон .xlsx для массовой загрузки студентов.',
    parameters=[AUTH_HEADER_PARAM],
    responses={
        200: OpenApiResponse(
            response=OpenApiTypes.BINARY,
            description='Файл .xlsx-шаблона (students_import_template.xlsx).',
        ),
        **ADMIN_RESPONSES,
    },
)
@api_view(['GET'])
@admin_required
def admin_bulk_import_template(request):
    """Скачать шаблон Excel для массовой загрузки студентов."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Студенты'

    headers = ['ФИО', 'Email', 'Телефон', 'Дата рождения', 'Группа']
    ws.append(headers)
    _style_header(ws, len(headers))

    example_fill = PatternFill('solid', fgColor='EBF3FB')
    ws.append(['Иванов Иван Иванович', 'ivan@example.com', '+992991234567', '2003-05-20', 'CS-101'])
    for col in range(1, len(headers) + 1):
        ws.cell(row=2, column=col).fill = example_fill
        ws.cell(row=2, column=col).font = Font(italic=True, color='555555')

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="students_import_template.xlsx"'
    return response
