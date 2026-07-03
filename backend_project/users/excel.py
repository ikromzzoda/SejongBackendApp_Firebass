import io

from django.http import HttpResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_COL_ALIASES = {
    'fullname':      ['full name', 'фио', 'ф.и.о', 'ф.и.о.', 'имя', 'полное имя', 'name', 'имя фамилия'],
    'email':         ['email', 'почта', 'e-mail', 'эл. почта', 'электронная почта'],
    'phone_number':  ['phone number', 'phone_number', 'телефон', 'номер', 'номер телефона', 'моб', 'моб.', 'тел'],
    'date_of_birth': ['date of birth/생년월일', 'дата рождения', 'дата', 'birth', 'д.р.', 'день рождения'],
    'group':         ['group', 'группа', 'учебная группа', 'класс'],
}


def _detect_columns(header_row) -> dict:
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell.value is None:
            continue
        normalized = str(cell.value).strip().lower()
        for field, aliases in _COL_ALIASES.items():
            if normalized in aliases and field not in mapping:
                mapping[field] = idx
    return mapping


def _style_header(ws, col_count: int):
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _open_import_workbook(excel_file):
    """Открывает входной .xlsx и возвращает (rows, error).
    При ошибке rows = None, error — текст для ответа 400.
    """
    try:
        wb_in = load_workbook(excel_file, data_only=True)
    except Exception:
        return None, 'Не удалось открыть файл. Убедитесь что это корректный .xlsx'

    ws_in = wb_in.active
    rows  = list(ws_in.iter_rows())
    if len(rows) < 2:
        return None, 'Файл пустой или содержит только заголовок'

    return rows, None


def _build_import_results_xlsx(results) -> bytes:
    """Собирает .xlsx с результатами импорта (логины/пароли/ошибки)."""
    wb_out  = Workbook()
    ws_out  = wb_out.active
    ws_out.title = 'Результаты импорта'
    ws_out.row_dimensions[1].height = 20

    headers = ['№', 'ФИО', 'Email', 'Телефон', 'Группа', 'Username', 'Password', 'Статус', 'Примечание']
    ws_out.append(headers)
    _style_header(ws_out, len(headers))

    green_fill = PatternFill('solid', fgColor='E2EFDA')
    red_fill   = PatternFill('solid', fgColor='FFDDC1')

    for i, (fullname, email, phone, group, username, password, status_text, note) in enumerate(results, start=1):
        ws_out.append([i, fullname, email, phone, group, username, password, status_text, note])
        row_fill = green_fill if status_text == 'Успешно' else red_fill
        for col in range(1, len(headers) + 1):
            ws_out.cell(row=i + 1, column=col).fill      = row_fill
            ws_out.cell(row=i + 1, column=col).alignment = Alignment(vertical='center')

    for row in ws_out.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.font = Font(bold=True)

    _auto_width(ws_out)

    success_count = sum(1 for r in results if r[6] == 'Успешно')
    error_count   = len(results) - success_count

    ws_out.append([])
    ws_out.append(['', f'Итого: {len(results)} строк | Успешно: {success_count} | Ошибок: {error_count}'])
    ws_out.cell(row=ws_out.max_row, column=2).font = Font(bold=True, size=11)

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output.getvalue()


def _build_import_template_xlsx() -> bytes:
    """Собирает .xlsx-шаблон для массовой загрузки студентов."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Студенты'

    headers = ['ФИО', 'Email', 'Телефон', 'Дата рождения', 'Группа']
    ws.append(headers)
    _style_header(ws, len(headers))

    example_fill = PatternFill('solid', fgColor='EBF3FB')
    ws.append(['Иванов Иван Иванович', 'ivan@example.com', '+992991234567', '2003-05-20', 'CS-101'])
    for col in range(1, len(headers) + 1):
        ws.cell(row=2, column=col).fill = example_fill
        ws.cell(row=2, column=col).font = Font(italic=True, color='555555')

    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _xlsx_response(content: bytes, filename: str) -> HttpResponse:
    """Оборачивает готовый .xlsx в HTTP-ответ с нужными заголовками."""
    response = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
